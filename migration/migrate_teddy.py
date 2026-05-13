"""
Migrate the Teddy 10U Apps Script data into TeamStats / Supabase.

Reads the local xlsx that the Apps Script app keeps in sync with the Google Sheet
(the same one at "C:\\Users\\mgrea\\OneDrive\\Desktop\\Spreadsheet for stats\\Teddy's 10U 2026.xlsx"
that build_stats.py generates and that the app writes to via the Sheets API).

Inserts into Supabase, mapping by player full name + game date so foreign keys line up.

Idempotent: if a team with the given slug already exists, the existing roster /
schedule rows are reused (matched by name + date) and game_log/lineups upsert
on conflict — so you can run this script multiple times safely.

Usage:
    py migrate_teddy.py <team-slug> <supabase-service-role-key>

Example:
    py migrate_teddy.py teddy-10u-2026 sb_secret_xxxxxxxxxxxxxxxxxxxxxx

The team must already exist (create it via the app first so RLS, ownership
and the trigger that adds you as owner all fire correctly).
"""

import sys, json, re, urllib.request, urllib.parse
import openpyxl
from datetime import datetime

# ====================================================================
# CONFIG
# ====================================================================
SUPABASE_URL = 'https://gqtewyseygisbcvwvyrg.supabase.co'
XLSX = r"C:\Users\mgrea\OneDrive\Desktop\Spreadsheet for stats\Teddy's 10U 2026.xlsx"

LAST_NAMES = {
    'Emily':   'Chase',
    'Nora':    'Buynicki',
    'Scarlett':'Larsen',
    'Maicey':  'Marte',
    'Lydia':   'Crane',
    'Marliey': 'Marte',
    'Katie':   'Greany',
    'Emerson': 'LaValley',
    'Brinn':   'Blascak',
    'Maisy':   'Call',
    'Reese':   'Humphrey',
    'Lyleese': 'Robinson-Meeker',
    'Raegan':  'Fennington',
}

# ====================================================================
# SUPABASE HELPERS (raw HTTP — no extra deps)
# ====================================================================
class Supa:
    def __init__(self, base, key):
        self.base = base.rstrip('/')
        self.key  = key
        self.headers = {
            'apikey': key,
            'Authorization': f'Bearer {key}',
            'Content-Type': 'application/json',
            'Prefer': 'return=representation',
        }

    def _req(self, method, path, body=None, headers=None):
        url = f'{self.base}/rest/v1/{path}'
        data = None
        if body is not None:
            data = json.dumps(body).encode('utf-8')
        req = urllib.request.Request(url, data=data, method=method)
        h = dict(self.headers)
        if headers: h.update(headers)
        for k, v in h.items():
            req.add_header(k, v)
        try:
            with urllib.request.urlopen(req) as resp:
                txt = resp.read().decode('utf-8')
                return json.loads(txt) if txt else None
        except urllib.error.HTTPError as e:
            err_body = e.read().decode('utf-8')
            print(f'HTTP {e.code} on {method} {path}: {err_body}', file=sys.stderr)
            raise

    def select(self, table, params=''):
        return self._req('GET', f'{table}{"?" + params if params else ""}')

    def insert(self, table, rows, *, on_conflict=None, return_min=False):
        path = table
        if on_conflict:
            path += f'?on_conflict={on_conflict}'
        h = {'Prefer': 'resolution=merge-duplicates,return=representation' if on_conflict
                       else 'return=representation'}
        if return_min:
            h['Prefer'] = h['Prefer'].replace('return=representation', 'return=minimal')
        return self._req('POST', path, body=rows, headers=h)

    def update(self, table, params, body):
        return self._req('PATCH', f'{table}?{params}', body=body)

# ====================================================================
# UTILITIES
# ====================================================================
def normalize_date(v):
    """Convert any of: '4/27', '4/27/2026', datetime, '2026-04-27' → 'YYYY-MM-DD'."""
    if v is None or v == '':
        return None
    if isinstance(v, datetime):
        return v.strftime('%Y-%m-%d')
    s = str(v).strip()
    m = re.match(r'^(\d{4})-(\d{2})-(\d{2})', s)
    if m: return f'{m.group(1)}-{m.group(2)}-{m.group(3)}'
    m = re.match(r'^(\d{1,2})/(\d{1,2})(?:/(\d{2,4}))?', s)
    if m:
        mo, da = int(m.group(1)), int(m.group(2))
        yr = int(m.group(3)) if m.group(3) else 2026
        if yr < 100: yr += 2000
        return f'{yr:04d}-{mo:02d}-{da:02d}'
    return None

def parse_full_name(full):
    """Given 'Emily Chase' or 'Emily', return (first, last)."""
    if not full: return (None, None)
    parts = full.strip().split(' ', 1)
    return (parts[0], parts[1] if len(parts) > 1 else '')

# ====================================================================
# MAIN
# ====================================================================
def main():
    if len(sys.argv) < 3:
        print(__doc__)
        sys.exit(1)
    slug = sys.argv[1]
    key  = sys.argv[2]
    supa = Supa(SUPABASE_URL, key)

    print(f'\n→ Looking up team "{slug}"...')
    teams = supa.select('teams', f'slug=eq.{urllib.parse.quote(slug)}&select=id,name')
    if not teams:
        print(f'  ✗ No team with slug "{slug}". Create it via the app first.')
        sys.exit(1)
    team = teams[0]
    team_id = team['id']
    print(f'  ✓ {team["name"]} ({team_id})')

    print(f'\n→ Opening {XLSX}...')
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    print(f'  ✓ Sheets: {wb.sheetnames}')

    # ----- 1. ROSTER -----
    print(f'\n→ Inserting roster...')
    # Use the canonical PLAYERS list from build_stats.py via LAST_NAMES
    roster_rows = []
    for i, (first, last) in enumerate(LAST_NAMES.items()):
        roster_rows.append({
            'team_id': team_id,
            'first_name': first,
            'last_name': last,
            'display_order': i,
        })
    # Skip if players already exist for this team
    existing = supa.select('players', f'team_id=eq.{team_id}&select=id,first_name,last_name')
    if existing:
        print(f'  • Roster already has {len(existing)} players — skipping insert; will reuse.')
        players = existing
    else:
        players = supa.insert('players', roster_rows)
        print(f'  ✓ Inserted {len(players)} players')
    player_id_by_name = {}      # full name "Emily Chase" → uuid
    player_id_by_first = {}     # "Emily" → uuid
    for p in players:
        full = f'{p["first_name"]} {p["last_name"] or ""}'.strip()
        player_id_by_name[full] = p['id']
        player_id_by_first[p['first_name']] = p['id']

    # ----- 2. SCHEDULE -----
    print(f'\n→ Inserting schedule...')
    sched_rows = []
    if 'Schedule' in wb.sheetnames:
        sh = wb['Schedule']
        for row in sh.iter_rows(min_row=2, values_only=True):
            date = normalize_date(row[0])
            opp  = row[1]
            ha   = (row[2] or 'home').lower()
            if not date or not opp: continue
            sched_rows.append({
                'team_id': team_id,
                'date': date,
                'opponent': str(opp),
                'home_away': ha if ha in ('home', 'away') else 'home',
            })
    if not sched_rows:
        # Fallback: hardcoded GAMES from build_stats.py
        GAMES = [
            ('4/27','vs. DSG','home'), ('4/29','vs. Amherst','home'),
            ('5/2','vs. Soho','home'), ('5/5','vs. E.Hamp','home'),
            ('5/7','vs. DSG','home'), ('5/9','vs. Westfield','home'),
            ('5/11','vs. Northampton','home'), ('5/13','@ Southampton','away'),
            ('5/16','vs. Westfield','home'), ('5/18','vs. Florence','home'),
            ('5/20','vs. Menard','home'), ('5/27','@ Fleury','away'),
        ]
        sched_rows = [{
            'team_id': team_id, 'date': normalize_date(d), 'opponent': o, 'home_away': ha
        } for d, o, ha in GAMES]
    existing_games = supa.select('games', f'team_id=eq.{team_id}&select=id,date,opponent')
    if existing_games:
        print(f'  • Schedule already has {len(existing_games)} games — skipping insert; will reuse.')
        games = existing_games
    else:
        games = supa.insert('games', sched_rows)
        print(f'  ✓ Inserted {len(games)} games')
    game_id_by_date = {g['date']: g['id'] for g in games}

    # ----- 3. GAME LOG -----
    print(f'\n→ Inserting game log...')
    if 'Game Log' not in wb.sheetnames:
        print('  • No Game Log sheet, skipping.')
    else:
        sh = wb['Game Log']
        # Column layout (matches Apps Script):
        # A=Date B=Opp C=Player D=AB E=R F=H G=1B H=2B I=BB J=HBP K=RBI L=SB M=Notes N=HR
        gl_rows = []
        for row in sh.iter_rows(min_row=4, values_only=True):
            date = normalize_date(row[0])
            player_full = row[2]
            if not date or not player_full: continue
            pid = player_id_by_name.get(player_full) or player_id_by_first.get(str(player_full).split()[0])
            gid = game_id_by_date.get(date)
            if not pid or not gid:
                print(f'  ! Skip: {player_full} on {date} (player or game not found)')
                continue
            gl_rows.append({
                'team_id': team_id, 'game_id': gid, 'player_id': pid,
                'ab':  int(row[3] or 0), 'r':   int(row[4] or 0), 'h':   int(row[5] or 0),
                'b1':  int(row[6] or 0), 'b2':  int(row[7] or 0),
                'bb':  int(row[8] or 0), 'hbp': int(row[9] or 0),
                'rbi': int(row[10] or 0), 'sb': int(row[11] or 0),
                'notes': str(row[12]) if row[12] else None,
                'hr':  int(row[13] or 0) if len(row) > 13 and row[13] is not None else 0,
            })
        if gl_rows:
            # Upsert on (game_id, player_id) so re-runs overwrite cleanly
            result = supa.insert('game_log', gl_rows, on_conflict='game_id,player_id')
            print(f'  ✓ Upserted {len(gl_rows)} game-log rows')
        else:
            print('  • No game-log rows found.')

    # ----- 4. PLAYER PROFILES -----
    print(f'\n→ Updating player profiles...')
    if 'Player Profiles' not in wb.sheetnames:
        print('  • No Player Profiles sheet, skipping.')
    else:
        sh = wb['Player Profiles']
        # Header: Player Jersey Position Bats Throws Height Age DOB Favorite Color Bio
        updated = 0
        for row in sh.iter_rows(min_row=2, values_only=True):
            full = row[0]
            if not full: continue
            pid = player_id_by_name.get(full) or player_id_by_first.get(str(full).split()[0])
            if not pid: continue
            payload = {
                'jersey':         str(row[1]) if row[1] is not None else None,
                'position':       str(row[2]) if row[2] else None,
                'bats':           (str(row[3]).strip().upper() if row[3] else '') or '',
                'throws':         (str(row[4]).strip().upper() if row[4] else '') or '',
                'height':         str(row[5]) if row[5] else None,
                'age':            str(row[6]) if row[6] is not None else None,
                'dob':            str(row[7]) if row[7] else None,
                'favorite_color': str(row[8]) if row[8] else None,
                'bio':            str(row[9]) if row[9] else None,
            }
            # Sanitize bats / throws
            if payload['bats']   not in ('L','R','S',''): payload['bats']   = ''
            if payload['throws'] not in ('L','R',''):     payload['throws'] = ''
            supa.update('players', f'id=eq.{pid}', payload)
            updated += 1
        print(f'  ✓ Updated {updated} player profiles')

    # ----- 5. LINEUPS -----
    print(f'\n→ Inserting lineups...')
    if 'Lineups' not in wb.sheetnames:
        print('  • No Lineups sheet, skipping.')
    else:
        sh = wb['Lineups']
        # Header: Date Order Positions Notes Updated
        ln_rows = []
        for row in sh.iter_rows(min_row=2, values_only=True):
            date = normalize_date(row[0])
            gid = game_id_by_date.get(date)
            if not gid: continue
            try: order_full = json.loads(row[1] or '[]')
            except Exception: order_full = []
            try: positions_full = json.loads(row[2] or '{}')
            except Exception: positions_full = {}
            # Convert name-keyed to UUID-keyed
            order_ids = []
            for n in order_full:
                pid = player_id_by_name.get(n) or player_id_by_first.get(str(n).split()[0])
                if pid: order_ids.append(pid)
            positions_by_id = {}
            for n, arr in positions_full.items():
                pid = player_id_by_name.get(n) or player_id_by_first.get(str(n).split()[0])
                if pid: positions_by_id[pid] = arr
            ln_rows.append({
                'game_id': gid, 'team_id': team_id,
                'batting_order': order_ids,
                'positions':     positions_by_id,
                'notes':         str(row[3]) if row[3] else None,
            })
        if ln_rows:
            result = supa.insert('lineups', ln_rows, on_conflict='game_id')
            print(f'  ✓ Upserted {len(ln_rows)} lineups')
        else:
            print('  • No lineup rows found.')

    print(f'\n✅ Migration complete. Open https://teamstats.mgreany.workers.dev/#/t/{slug}')

if __name__ == '__main__':
    main()
